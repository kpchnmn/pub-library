from os import path
import boto3
import configparser
import openpyxl as px
from openpyxl.styles import Font
import aws
from tkinter import Tk ,messagebox

SECTION1 = 'excel'
SECTION2 = 'aws configure'

FILENAME = 'filename'
FONT = 'font'
VPC_TITLE = 'vpc_title'
SUBNET_TITLE = 'subnet_title'
ROUTE_TITLE = 'route_title'
IGW_TITLE = 'igw_title'
DHCP_TITLE = 'dhcp_title'
EIP_TITLE = 'eip_title'
NGW_TITLE = 'ngw_title'
NACL_TITLE = 'nacl_title'
SG_TITLE = 'sg_title'
CGW_TITLE = 'cgw_title'
VGW_TITLE = 'vgw_title'
CVEP_TITLE = 'cvep_title'
EC2_TITLE = 'ec2_title'
EBS_TITLE = 'ebs_title'
ENI_TITLE = 'eni_title'

FILENAME_EXAMPLE = 'aws_parameter.xlsx'
FONT_EXAMPLE = 'メイリオ'
VPC_TITLE_EXAMPLE = 'VPC'
SUBNET_TITLE_EXAMPLE = 'サブネット'
ROUTE_TITLE_EXAMPLE = 'ルートテーブル'
IGW_TITLE_EXAMPLE = 'インターネットゲートウェイ'
DHCP_TITLE_EXAMPLE = 'DHCPオプションセット'
EIP_TITLE_EXAPMLE = 'ElasticIp'
NGW_TITLE_EXAMPLE = 'NATゲートウェイ'
NACL_TITLE_EXAMPLE = 'ネットワークACL'
SG_TITLE_EXAMPLE = 'セキュリティグループ'
CGW_TITLE_EXAMPLE = 'カスタマーゲートウェイ'
VGW_TITLE_EXAMPLE = '仮想プライベートゲートウェイ'
CVEP_TITLE_EXAMPLE = 'クライアントVPNエンドポイント'
EC2_TITLE_EXAMPLE = 'EC2インスタンス'
EBS_TITLE_EXAMPLE = 'EC2ボリューム'
ENI_TITLE_EXAMPLE = 'ネットワークインターフェース'

DEFAULT= 'default'
AWS_ACCESS_KEY_ID = 'aws_access_key_id'
AWS_SECRET_ACCESS_KEY = 'aws_secret_access_key'
REGION_NAME = 'region_name'
REGION_EXAMPLE = 'ap-northeast-1'
AWS_PATH = path.expanduser('~/.aws/credentials')

TITLE = 'AWSパラメーターシート作成'
ERROR_MSG = 'アプリケーションエラーが発生しました。'

CONFIG = 'setting.ini'
ENCODING = 'UTF-8'

STATUS = True

def get_aws_configure(path):
    config.read(path,ENCODING)
    return config.get(DEFAULT,AWS_ACCESS_KEY_ID), config.get(DEFAULT,AWS_SECRET_ACCESS_KEY)

def create_setting():
    try:
        config.add_section(SECTION1)
    except:
        pass
    config.set(SECTION1,FILENAME,FILENAME_EXAMPLE)
    config.set(SECTION1,FONT,FONT_EXAMPLE)
    config.set(SECTION1,VPC_TITLE,VPC_TITLE_EXAMPLE)
    config.set(SECTION1,SUBNET_TITLE,SUBNET_TITLE_EXAMPLE)
    config.set(SECTION1,ROUTE_TITLE,ROUTE_TITLE_EXAMPLE)
    config.set(SECTION1,IGW_TITLE,IGW_TITLE_EXAMPLE)
    config.set(SECTION1,DHCP_TITLE,DHCP_TITLE_EXAMPLE)
    config.set(SECTION1,EIP_TITLE,EIP_TITLE_EXAPMLE)
    config.set(SECTION1,NGW_TITLE,NGW_TITLE_EXAMPLE)
    config.set(SECTION1,NACL_TITLE,NACL_TITLE_EXAMPLE)
    config.set(SECTION1,SG_TITLE,SG_TITLE_EXAMPLE)
    config.set(SECTION1,CGW_TITLE,CGW_TITLE_EXAMPLE)
    config.set(SECTION1,VGW_TITLE,VGW_TITLE_EXAMPLE)
    config.set(SECTION1,CVEP_TITLE,CVEP_TITLE_EXAMPLE)
    config.set(SECTION1,EC2_TITLE,EC2_TITLE_EXAMPLE)
    config.set(SECTION1,EBS_TITLE,EBS_TITLE_EXAMPLE)
    config.set(SECTION1,ENI_TITLE,ENI_TITLE_EXAMPLE)

    try:
        config.add_section(SECTION2)
    except:
        pass
    try:
        config.set(SECTION2,AWS_ACCESS_KEY_ID, get_aws_configure(AWS_PATH)[0] )
        config.set(SECTION2,AWS_SECRET_ACCESS_KEY, get_aws_configure(AWS_PATH)[1])
        config.set(SECTION2,REGION_NAME,REGION_EXAMPLE)
    except:
        quit()

    with open(CONFIG, 'w',encoding=ENCODING) as file:
        config.write(file)

def read_config():
    global filename
    global font
    global vpc_title
    global subnet_title
    global route_title
    global igw_title
    global dhcp_title
    global eip_title
    global ngw_title
    global nacl_title
    global sg_title
    global cgw_title
    global vgw_title
    global cvep_title
    global ec2_title
    global ebs_title
    global eni_title

    global aws_access_key_id
    global aws_secret_access_key
    global region_name

    filename = config.get(SECTION1,FILENAME)
    font = config.get(SECTION1,FONT)
    vpc_title = config.get(SECTION1,VPC_TITLE)
    subnet_title = config.get(SECTION1,SUBNET_TITLE)
    route_title = config.get(SECTION1,ROUTE_TITLE)
    igw_title = config.get(SECTION1,IGW_TITLE)
    dhcp_title = config.get(SECTION1, DHCP_TITLE)
    eip_title = config.get(SECTION1,EIP_TITLE)
    ngw_title = config.get(SECTION1,NGW_TITLE)
    nacl_title = config.get(SECTION1,NACL_TITLE)
    sg_title = config.get(SECTION1,SG_TITLE)
    cgw_title = config.get(SECTION1,CGW_TITLE)
    vgw_title = config.get(SECTION1,VGW_TITLE)
    cvep_title = config.get(SECTION1,CVEP_TITLE)
    ec2_title = config.get(SECTION1,EC2_TITLE)
    ebs_title = config.get(SECTION1,EBS_TITLE)
    eni_title = config.get(SECTION1,ENI_TITLE)

    aws_access_key_id = config.get(SECTION2,AWS_ACCESS_KEY_ID)
    aws_secret_access_key = config.get(SECTION2,AWS_SECRET_ACCESS_KEY)
    region_name = config.get(SECTION2,REGION_NAME)

config = configparser.ConfigParser()

try:
    config.read(CONFIG,ENCODING)
    read_config()
except:
    create_setting()
    read_config()


vpc_row = ['Nameタグ','VPC ID','CidrBlock','DNS解決','DNSホスト名','DHCPオプションセット','テナンシー','デフォルトVPC']
subnet_row = ['Nameタグ','サブネットID','CidrBlock','AvailabilityZone','デフォルトのサブネット','パブリックIPv4アドレスの自動割り当て','VPC ID']
route_row = ['Nameタグ','ルートテーブルID','VPC ID','ルート伝播','メイン','関連付けサブネットID','送信先','ターゲット']
igw_row = ['Nameタグ','IGW ID','アタッチ済み VPC ID']
dhcp_row = ['Nameタグ','DHCPオプションセットID','domain-name','domain-name-servers']
eip_row = ['Nameタグ','プライベートIP','パブリックIP','割り当てID','関連ID','インスタンスID','ネットワークインターフェースID']
ngw_row = ['Nameタグ','NAT ゲートウェイ ID','VPC ID','サブネットID','プライベートIP','パブリックIP','ネットワークインターフェースID']
nacl_row = ['Nameタグ','ネットワーク ACL ID','デフォルト','VPC ID','関連付けサブネットID','インバウンド/ルール','プロトコル','送信元','許可/拒否','アウトバウンド/ルール','プロトコル','送信元','許可/拒否']
sg_row = ['Nameタグ','グループID','グループ名','説明','VPC ID','インバウンド/プロトコル','ポート範囲','ソース','説明','アウトバウンド/プロトコル','ポート範囲','送信先','説明']
cgw_row = ['Nameタグ','CGW ID','IPアドレス','BGP ASN','タイプ']
vgw_row = ['Nameタグ','VGW ID','状態','VPC ID','ASN (Amazon 側)','タイプ']
cvep_row = ['Nameタグ','CVEP ID','説明','クライアント IPv4 CIDR','DNS 名','トランスポートプロトコル','スプリットトンネル','認証タイプ','証明書 ARN','タイプ','CloudWatch ロググループ','CloudWatch ログストリーム','サブネットID','宛先のCIDR']
ec2_row = ['Nameタグ','インスタンスID','インスタンスタイプ','パブリックIP','パブリックDNS','プライベートIP','プライベートDNS','AMI ID','キーペア名','AvailabilityZone','IAMロール','VPC ID','サブネット ID','セキュリティグループ','ボリューム ID','ネットワークインターフェースID']
ebs_row = ['Nameタグ','ボリュームID','インスタンスID','サイズ','ボリュームタイプ','IOPS','暗号化','AvailabilityZone']
eni_row = ['Nameタグ','ネットワークインターフェイス ID','VPC ID','サブネット ID','アベイラビリティーゾーン','プライベートIPアドレス','MAC アドレス','説明','アタッチメント ID','送信元/送信先 チェック']

ec2 = boto3.client('ec2',
                aws_access_key_id=aws_access_key_id,
                aws_secret_access_key=aws_secret_access_key,
                region_name=region_name)

def create_file(filename):
    wb = px.Workbook()
    wb.save(filename)

def delete_sheet(filename):
    wb = px.load_workbook(filename)
    try:
        del wb['Sheet']
    except:
        pass
    wb.save(filename)

def main():
    root = Tk()
    root.withdraw()
    msg = TITLE +'を実行します。'
    STATUS = messagebox.askokcancel(TITLE,msg)
    if STATUS:
        try:
            px.load_workbook(filename)
        except:
            STATUS = False
            create_file(filename)
            STATUS = True
        try:
            aws.vpc.main(filename)
            aws.subnet.main(filename)
            aws.route.main(filename)
            aws.igw.main(filename)
            aws.dhcp.main(filename)
            aws.eip.main(filename)
            aws.ngw.main(filename)
            aws.nacl.main(filename)
            aws.sg.main(filename)
            aws.cgw.main(filename)
            aws.vgw.main(filename)
            #aws.cvep.main(filename)
            aws.ec2.main(filename)
            aws.ebs.main(filename)
            aws.eni.main(filename)
            delete_sheet(filename)
        except:
            STATUS = False
        if STATUS:
            msg = filename + 'を作成しました。'
            messagebox.showinfo(TITLE, msg)
        else:
            messagebox.showerror(TITLE,ERROR_MSG)
    else:
        msg = '実行をキャンセルしました。'
        messagebox.showinfo(TITLE,msg)

if __name__ == "__main__":
    main()
